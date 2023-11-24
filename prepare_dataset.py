from openpyxl import load_workbook, Workbook

if __name__ == '__main__':
    data = load_workbook('dataset/unprepared_dataset/BRICS Economic Data.xlsx')
    sheet = data['Data']

    countries = []
    for row in range(2, sheet.max_row):
        cell = sheet[row][1]
        if cell.value and cell.value not in countries:
            countries.append(cell.value)

    country_books = [Workbook() for _ in countries]
    column_countries = [2 for _ in countries]
    for book in country_books:
        current_sheet = book.active
        for column in range(1, sheet.max_column + 2):
            current_sheet.cell(column, 1).value = sheet.cell(1, column).value

    for row in range(2, sheet.max_row+2):
        if sheet[row][1].value in countries:
            country_index = countries.index(sheet[row][1].value)
            current_book = country_books[country_index]
            current_sheet = current_book.active
            for column in range(1, sheet.max_column+2):
                current_sheet.cell(column, column_countries[country_index]).value = sheet.cell(row, column).value
            column_countries[country_index] += 1

    for country in range(len(countries)):
        country_books[country].save(f'{countries[country]}_economy.xlsx')
